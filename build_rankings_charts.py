"""Add native charts to every sheet of RANKINGS_SUMMARY.xlsx.

openpyxl cannot author PivotTables from scratch, so this builds equivalent
native charts directly off each sheet's Excel Table, plus a small helper
"Division Summary" block (computed with AVERAGEIF/MAXIF/COUNTIF over the
table) that drives the per-division charts. Charts update live with the data.

Charts per sheet:
  * Overall sheets (4 score cols / 3 score cols): metric-disagreement scatter
    (z vs z), Fights-vs-Score scatter, and a per-division average bar.
  * Divisional sheets: per-division average + max bar, and a division depth
    (fighter count) bar.
"""
import openpyxl
from openpyxl.chart import ScatterChart, BarChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

XLSX = "RANKINGS_SUMMARY.xlsx"

HDR_ROW = 8
DATA_ROW = 9


def col_map(ws):
    return {ws.cell(row=HDR_ROW, column=c).value: c
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=HDR_ROW, column=c).value}


def last_data_row(ws, fighter_col):
    r = DATA_ROW
    while ws.cell(row=r, column=fighter_col).value not in (None, ""):
        r += 1
    return r - 1


def divisions_in(ws, div_col, r0, r1):
    seen = []
    for r in range(r0, r1 + 1):
        v = ws.cell(row=r, column=div_col).value
        if v and v not in seen:
            seen.append(v)
    return seen


def add_division_summary(ws, cm, r0, r1, score_label, anchor_col):
    """Write a Division | Avg | Max | Count helper block starting at
    (HDR_ROW, anchor_col). Returns (header_row, first_row, last_row, cols)."""
    divs = divisions_in(ws, cm["Division"], r0, r1)
    score_letter = get_column_letter(cm[score_label])
    div_letter = get_column_letter(cm["Division"])
    score_rng = f"${score_letter}${r0}:${score_letter}${r1}"
    div_rng = f"${div_letter}${r0}:${div_letter}${r1}"

    a = anchor_col
    ws.cell(row=HDR_ROW, column=a, value="Division").font = Font(bold=True)
    ws.cell(row=HDR_ROW, column=a + 1, value="Avg Score").font = Font(bold=True)
    ws.cell(row=HDR_ROW, column=a + 2, value="Max Score").font = Font(bold=True)
    ws.cell(row=HDR_ROW, column=a + 3, value="Fighters").font = Font(bold=True)
    for i, d in enumerate(divs):
        r = HDR_ROW + 1 + i
        dcell = ws.cell(row=r, column=a, value=d)
        dref = f"${get_column_letter(a)}${r}"
        ws.cell(row=r, column=a + 1,
                value=f"=AVERAGEIF({div_rng},{dref},{score_rng})").number_format = "0.0"
        ws.cell(row=r, column=a + 2,
                value=f"=MAXIFS({score_rng},{div_rng},{dref})").number_format = "0.0"
        ws.cell(row=r, column=a + 3,
                value=f"=COUNTIF({div_rng},{dref})")
    for off, w in [(0, 22), (1, 12), (2, 12), (3, 10)]:
        ws.column_dimensions[get_column_letter(a + off)].width = w
    return HDR_ROW, HDR_ROW + 1, HDR_ROW + len(divs), a


def scatter(ws, title, x_col, y_col, r0, r1, x_title, y_title,
            anchor, trend=False):
    ch = ScatterChart()
    ch.title = title
    ch.x_axis.title = x_title
    ch.y_axis.title = y_title
    ch.style = 13
    ch.height = 8
    ch.width = 13
    xref = Reference(ws, min_col=x_col, min_row=r0, max_row=r1)
    yref = Reference(ws, min_col=y_col, min_row=r0, max_row=r1)
    s = Series(yref, xref, title=y_title)
    s.marker.symbol = "circle"
    s.marker.size = 5
    s.graphicalProperties.line.noFill = True
    if trend:
        s.trendline = Trendline(trendlineType="linear")
    ch.series.append(s)
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ws.add_chart(ch, anchor)


def bar(ws, title, cat_col, val_cols, cat_r0, cat_r1, anchor,
        y_title="Score", hdr_row=HDR_ROW):
    ch = BarChart()
    ch.type = "col"
    ch.title = title
    ch.y_axis.title = y_title
    ch.style = 10
    ch.height = 8
    ch.width = 16
    cats = Reference(ws, min_col=cat_col, min_row=cat_r0, max_row=cat_r1)
    for vc in val_cols:
        data = Reference(ws, min_col=vc, min_row=hdr_row, max_row=cat_r1)
        ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ws.add_chart(ch, anchor)


def overall_sheet(ws, score_labels, z_pairs):
    """score_labels: ordered score columns. z_pairs: list of (zx,zy) header
    names for disagreement scatters."""
    cm = col_map(ws)
    r1 = last_data_row(ws, cm["Fighter"])
    r0 = DATA_ROW

    # helper block placed a couple columns past the table
    anchor_col = max(cm.values()) + 2
    h, s0, s1, a = add_division_summary(ws, cm, r0, r1, score_labels[0],
                                        anchor_col)

    # charts stacked below the helper block
    chart_anchor_col = get_column_letter(a)
    row_cursor = s1 + 3

    # 1. metric disagreement scatter (first z-pair)
    zx, zy = z_pairs[0]
    scatter(ws, f"Metric Disagreement: {zx} vs {zy}",
            cm[zx], cm[zy], r0, r1, zx, zy,
            f"{chart_anchor_col}{row_cursor}")
    row_cursor += 17

    # 2. fights vs headline score
    scatter(ws, "Fights vs WHR Score",
            cm["Fights"], cm[score_labels[0]], r0, r1,
            "Fights", score_labels[0],
            f"{chart_anchor_col}{row_cursor}", trend=True)
    row_cursor += 17

    # 3. per-division average + max bar
    bar(ws, "Division Strength (Avg & Max Score)",
        a, [a + 1, a + 2], s0, s1,
        f"{chart_anchor_col}{row_cursor}")


def divisional_sheet(ws, score_label="WHR Score"):
    cm = col_map(ws)
    r1 = last_data_row(ws, cm["Fighter"])
    r0 = DATA_ROW
    anchor_col = max(cm.values()) + 2
    h, s0, s1, a = add_division_summary(ws, cm, r0, r1, score_label,
                                        anchor_col)
    chart_anchor_col = get_column_letter(a)
    row_cursor = s1 + 3

    # 1. per-division avg & max
    bar(ws, "Division Strength (Avg & Max WHR)",
        a, [a + 1, a + 2], s0, s1,
        f"{chart_anchor_col}{row_cursor}")
    row_cursor += 17

    # 2. division depth (fighter count)
    bar(ws, "Division Depth (Fighter Count)",
        a, [a + 3], s0, s1,
        f"{chart_anchor_col}{row_cursor}", y_title="Fighters")
    row_cursor += 17

    # 3. fights vs score scatter
    scatter(ws, "Fights vs WHR Score",
            cm["Fights"], cm[score_label], r0, r1,
            "Fights", score_label,
            f"{chart_anchor_col}{row_cursor}", trend=True)


def main():
    wb = openpyxl.load_workbook(XLSX)

    # model sheet — 4 score columns, custom z header names
    overall_sheet(
        wb["MEN -WHR 10 yr"],
        ["WHR Score", "WHR I+P Score", "Method I+P Score", "Canonical Score"],
        z_pairs=[("z-Score WHR", "z-Method I+P")],
    )
    overall_sheet(
        wb["MEN -WHR 5 yr"],
        ["WHR Score", "WHR I+P Score", "Method I+P Score"],
        z_pairs=[("z WHR Score", "z Method I+P Score")],
    )
    overall_sheet(
        wb["WOMEN -Overall 10 yr"],
        ["WHR Score", "WHR I+P Score", "Method I+P Score"],
        z_pairs=[("z WHR Score", "z Method I+P Score")],
    )
    overall_sheet(
        wb["WOMEN -Overall 5 yr"],
        ["WHR Score", "WHR I+P Score", "Method I+P Score"],
        z_pairs=[("z WHR Score", "z Method I+P Score")],
    )
    divisional_sheet(wb["MEN -FlyW WHR"])
    divisional_sheet(wb["WOMEN -Divisional WHR"])

    wb.save(XLSX)
    print("Charts added to all sheets. Saved", XLSX)


if __name__ == "__main__":
    main()
