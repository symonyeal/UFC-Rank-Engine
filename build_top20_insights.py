"""Fix rank direction on the model sheet and add Top-20 insight charts to the
two men's overall sheets ('MEN -WHR 10 yr', 'MEN -WHR 5 yr').

Consistency fix
---------------
The model sheet's RANK formulas used ascending order (best fighter = highest
rank number). Flip every RANK(...,1) to RANK(...,0) so rank 1 = best, matching
the rebuilt sheets and the markdown ordering.

Top-20 insight block
--------------------
Below each sheet's Division Summary helper, write a "Top 20 by WHR" block:
ranks the table by the headline WHR score, pulls each fighter's per-metric
z-scores and per-metric ranks, and derives:
  * Avg Rank   - mean rank across all metrics (lower = stronger overall)
  * Rank Spread - max rank minus min rank across metrics (higher = the
    metrics disagree most about this fighter -> anomaly)
  * Z Spread   - max z minus min z across metrics

Charts:
  1. Top 20 - z-score profile by metric (clustered bar): flat bars = a
     fighter every metric agrees on; one tall/short bar = metric-specific.
  2. Top 20 - rank under each metric (line): crossing lines expose where a
     metric promotes/demotes a fighter vs the others.
  3. Consistency vs peak (scatter): X = Avg Rank, Y = Rank Spread. Lower-left
     = elite and consistent; high Y = anomaly the metrics disagree on.
"""
import openpyxl
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

XLSX = "RANKINGS_SUMMARY.xlsx"
HDR_ROW = 8
DATA_ROW = 9
TOP_N = 20


def table_col_map(ws, table_ref):
    """Header->col index, restricted to the Excel Table's own column span so
    helper-block headers with duplicate text don't collide."""
    first, last = table_ref.split(":")
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    c0 = column_index_from_string(coordinate_from_string(first)[0])
    c1 = column_index_from_string(coordinate_from_string(last)[0])
    return {ws.cell(row=HDR_ROW, column=c).value: c for c in range(c0, c1 + 1)}


def last_data_row(ws, fighter_col):
    r = DATA_ROW
    while ws.cell(row=r, column=fighter_col).value not in (None, ""):
        r += 1
    return r - 1


def fix_rank_direction(ws):
    """Flip RANK(...,1) -> RANK(...,0) everywhere on the sheet."""
    n = 0
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("=") and "RANK(" in v:
                if v.rstrip().endswith(",1)"):
                    c.value = v.rstrip()[:-3] + ",0)"
                    n += 1
    return n


def find_helper_end(ws):
    """Row where the existing Division Summary helper block ends."""
    # locate the 'Division' header of the helper (a header cell past col R)
    for c in range(15, ws.max_column + 2):
        if (ws.cell(row=HDR_ROW, column=c).value == "Division"
                and ws.cell(row=HDR_ROW, column=c + 1).value == "Avg Score"):
            r = HDR_ROW + 1
            while ws.cell(row=r, column=c).value not in (None, ""):
                r += 1
            return r - 1
    return HDR_ROW


def build_insight_block(ws, table_name, score_labels, start_col):
    """Write Top-20 insight table starting at (HDR_ROW, start_col).
    Returns dict of useful (col, row) refs for charting."""
    cm = table_col_map(ws, ws.tables[table_name].ref)
    r1 = last_data_row(ws, cm["Fighter"])
    fighter_letter = get_column_letter(cm["Fighter"])
    whr_letter = get_column_letter(cm[score_labels[0]])
    tbl_score = f"{table_name}[{score_labels[0]}]"

    a = start_col
    title_row = HDR_ROW - 1  # mirror the sheet's title placement style
    ws.cell(row=title_row, column=a,
            value=f"Top {TOP_N} Men — Metric Insight").font = Font(size=14, bold=True)

    headers = (["Rank", "Fighter"] + [f"z {s}" for s in score_labels]
               + [f"R {s}" for s in score_labels]
               + ["Avg Rank", "Rank Spread", "Z Spread"])
    hdr_fill = PatternFill("solid", fgColor="305496")
    for j, h in enumerate(headers):
        cell = ws.cell(row=HDR_ROW, column=a + j, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    col = {h: a + j for j, h in enumerate(headers)}
    nz = len(score_labels)

    for i in range(TOP_N):
        r = DATA_ROW + i
        rank_pos = i + 1
        # rank position label
        ws.cell(row=r, column=col["Rank"], value=rank_pos)
        # fighter ranked rank_pos by headline WHR (descending)
        f_formula = (
            f"=INDEX({table_name}[Fighter],"
            f"MATCH(LARGE({tbl_score},{rank_pos}),{tbl_score},0))"
        )
        ws.cell(row=r, column=col["Fighter"], value=f_formula)
        fcell = f"${get_column_letter(col['Fighter'])}${r}"

        # per-metric z-scores and ranks via lookup on Fighter
        for s in score_labels:
            zc = ws.cell(row=r, column=col[f"z {s}"])
            zc.value = (
                f"=INDEX({table_name}[{_resolve_zrank(table_name,s,'z')}],"
                f"MATCH({fcell},{table_name}[Fighter],0))"
            )
            zc.number_format = "0.00"
            rc = ws.cell(row=r, column=col[f"R {s}"])
            rc.value = (
                f"=INDEX({table_name}[{_resolve_zrank(table_name,s,'rank')}],"
                f"MATCH({fcell},{table_name}[Fighter],0))"
            )

        # derived consistency metrics over this row's R/z cells
        r_first = get_column_letter(col[f"R {score_labels[0]}"])
        r_last = get_column_letter(col[f"R {score_labels[-1]}"])
        z_first = get_column_letter(col[f"z {score_labels[0]}"])
        z_last = get_column_letter(col[f"z {score_labels[-1]}"])
        ws.cell(row=r, column=col["Avg Rank"],
                value=f"=AVERAGE(${r_first}{r}:${r_last}{r})").number_format = "0.0"
        ws.cell(row=r, column=col["Rank Spread"],
                value=f"=MAX(${r_first}{r}:${r_last}{r})-MIN(${r_first}{r}:${r_last}{r})")
        ws.cell(row=r, column=col["Z Spread"],
                value=f"=MAX(${z_first}{r}:${z_last}{r})-MIN(${z_first}{r}:${z_last}{r})"
                ).number_format = "0.00"

    # column widths
    ws.column_dimensions[get_column_letter(col["Rank"])].width = 6
    ws.column_dimensions[get_column_letter(col["Fighter"])].width = 22
    for s in score_labels:
        ws.column_dimensions[get_column_letter(col[f"z {s}"])].width = 11
        ws.column_dimensions[get_column_letter(col[f"R {s}"])].width = 9
    for h in ("Avg Rank", "Rank Spread", "Z Spread"):
        ws.column_dimensions[get_column_letter(col[h])].width = 12

    end_row = DATA_ROW + TOP_N - 1
    return col, end_row


# The rebuilt sheets and the model sheet name their z/rank columns slightly
# differently; resolve the real header text used inside each table.
_ZRANK_CACHE = {}


def _resolve_zrank(table_name, score_label, kind):
    """kind: 'z' or 'rank'. Return the actual header text used in this table."""
    # model sheet (Table26) uses bespoke names; rebuilt sheets use 'z <label>'
    if table_name == "Table26":
        zmap = {
            "WHR Score": "z-Score WHR",
            "WHR I+P Score": "z-Score WHR I+P",
            "Method I+P Score": "z-Method I+P",
            "Canonical Score": "z-Canonical Glicko2",
        }
        rmap = {
            "WHR Score": "Rank WHR",
            "WHR I+P Score": "Rank WHR I+P",
            "Method I+P Score": "Rank Method I+P",
            "Canonical Score": "Rank Canonical",
        }
    else:
        zmap = {s: f"z {s}" for s in (
            "WHR Score", "WHR I+P Score", "Method I+P Score", "Canonical Score")}
        rmap = {s: f"Rank {s}" for s in (
            "WHR Score", "WHR I+P Score", "Method I+P Score", "Canonical Score")}
    return (zmap if kind == "z" else rmap)[score_label]


def add_insight_charts(ws, col, end_row, score_labels):
    anchor_col_idx = col["Rank"]
    chart_col = get_column_letter(anchor_col_idx)
    cursor = end_row + 3

    fighter_col = col["Fighter"]
    r0, r1 = DATA_ROW, end_row

    # 1. z-score profile by metric (clustered bar)
    ch = BarChart()
    ch.type = "col"
    ch.title = f"Top {TOP_N}: z-Score Profile by Metric"
    ch.y_axis.title = "z-score (vs field)"
    ch.style = 10
    ch.height = 9
    ch.width = 22
    cats = Reference(ws, min_col=fighter_col, min_row=r0, max_row=r1)
    for s in score_labels:
        data = Reference(ws, min_col=col[f"z {s}"], min_row=HDR_ROW, max_row=r1)
        ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ws.add_chart(ch, f"{chart_col}{cursor}")
    cursor += 19

    # 2. rank under each metric (line)
    ch = LineChart()
    ch.title = f"Top {TOP_N}: Rank Under Each Metric (1 = best)"
    ch.y_axis.title = "Rank"
    ch.style = 12
    ch.height = 9
    ch.width = 22
    ch.y_axis.scaling.orientation = "maxMin"  # rank 1 at top
    for s in score_labels:
        data = Reference(ws, min_col=col[f"R {s}"], min_row=HDR_ROW, max_row=r1)
        ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ws.add_chart(ch, f"{chart_col}{cursor}")
    cursor += 19

    # 3. consistency vs peak (scatter): X = Avg Rank, Y = Rank Spread
    ch = ScatterChart()
    ch.title = f"Top {TOP_N}: Consistency vs Peak"
    ch.x_axis.title = "Avg Rank across metrics (lower = stronger)"
    ch.y_axis.title = "Rank Spread (higher = metrics disagree)"
    ch.style = 13
    ch.height = 11
    ch.width = 16
    xref = Reference(ws, min_col=col["Avg Rank"], min_row=r0, max_row=r1)
    yref = Reference(ws, min_col=col["Rank Spread"], min_row=r0, max_row=r1)
    s = Series(yref, xref, title="Top 20 fighter")
    s.marker.symbol = "circle"
    s.marker.size = 7
    s.graphicalProperties.line.noFill = True
    ch.series.append(s)
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ws.add_chart(ch, f"{chart_col}{cursor}")


def process(ws, table_name, score_labels):
    flipped = fix_rank_direction(ws)
    helper_end = find_helper_end(ws)
    # insight block starts a couple rows below helper, same start column as helper
    # find helper start col
    start_col = None
    for c in range(15, ws.max_column + 2):
        if (ws.cell(row=HDR_ROW, column=c).value == "Division"
                and ws.cell(row=HDR_ROW, column=c + 1).value == "Avg Score"):
            start_col = c
            break
    if start_col is None:
        start_col = ws.max_column + 2
    # place insight block to the RIGHT of the helper block (helper is 4 cols wide)
    insight_col = start_col + 6
    col, end_row = build_insight_block(
        ws, table_name, score_labels, insight_col)
    add_insight_charts(ws, col, end_row, score_labels)
    return flipped


def main():
    wb = openpyxl.load_workbook(XLSX)

    f1 = process(wb["MEN -WHR 10 yr"], "Table26",
                 ["WHR Score", "WHR I+P Score", "Method I+P Score",
                  "Canonical Score"])
    f2 = process(wb["MEN -WHR 5 yr"], "TableMen5yr",
                 ["WHR Score", "WHR I+P Score", "Method I+P Score"])

    # also flip rank direction on the remaining rebuilt sheets for consistency
    flips = {"MEN -WHR 10 yr": f1, "MEN -WHR 5 yr": f2}
    for name in ["MEN -FlyW WHR", "WOMEN -Overall 10 yr",
                 "WOMEN -Divisional WHR", "WOMEN -Overall 5 yr"]:
        flips[name] = fix_rank_direction(wb[name])

    wb.save(XLSX)
    print("Saved.", XLSX)
    for k, v in flips.items():
        print(f"  {k}: {v} RANK formulas flipped to descending")


if __name__ == "__main__":
    main()
