"""Rebuild RANKINGS_SUMMARY.xlsx secondary sheets to match the model sheet
('MEN -WHR 10 yr'): one merged Excel Table per sheet, gaps filled from the
markdown source, z-score + rank formulas appended.
"""
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

MD = "RANKINGS_SUMMARY.md"
XLSX = "RANKINGS_SUMMARY.xlsx"
RATINGS_PARQUET = "data/snapshots/2026-05-13/ratings_current.parquet"

# Markdown score-column label -> DB column suffix (prefix added per era).
# era prefix: "sustained_peak_headline_mu_" (10yr) / "five_year_peak_headline_mu_" (5yr)
DB_SUFFIX = {
    "WHR Score": "whr",
    "WHR I+P Score": "whr_integrity_performance",
    "Method I+P Score": "method_integrity_performance",
    "Canonical Score": "canonical",
}


def parse_md():
    section = None
    cur = None
    tables = {}
    for ln in open(MD, encoding="utf-8").read().split("\n"):
        if ln.startswith("## "):
            section = ln[3:].strip()
        elif ln.startswith("### "):
            sub = ln[4:].strip()
            cur = []
            tables[(section, sub)] = cur
        elif ln.startswith("|") and cur is not None:
            cells = [c.strip() for c in ln.strip("|").split("|")]
            if set("".join(cells)) <= set("-: "):
                continue
            cur.append(cells)
    # normalize: drop header row, return list of dict rows
    out = {}
    for (sec, sub), rows in tables.items():
        hdr = rows[0]
        data = []
        for r in rows[1:]:
            d = dict(zip(hdr, r))
            data.append(d)
        out[(sec, sub)] = (hdr, data)
    return out


def men_section(t):
    return next(k for k in t if k[0].startswith("MEN") and "Overall" in k[0])[0]


def main():
    t = parse_md()
    # resolve unicode-dash section names dynamically
    men_overall = next(k[0] for k in t if k[0].startswith("MEN") and "Overall" in k[0])
    men_div = next(k[0] for k in t if k[0].startswith("MEN") and "Divisional" in k[0])
    wom_overall = next(k[0] for k in t if k[0].startswith("WOMEN") and "Overall" in k[0])
    wom_div = next(k[0] for k in t if k[0].startswith("WOMEN") and "Divisional" in k[0])

    wb = openpyxl.load_workbook(XLSX)
    ratings = pd.read_parquet(RATINGS_PARQUET).set_index("fighter")

    # ---- helpers -------------------------------------------------------
    def db_score(fighter, out_label, era_prefix):
        """Look up a peak score from ratings_current; None if unavailable."""
        suffix = DB_SUFFIX.get(out_label)
        if suffix is None or fighter not in ratings.index:
            return None
        col = era_prefix + suffix
        if col not in ratings.columns:
            return None
        v = ratings.at[fighter, col]
        if pd.isna(v):
            return None
        return round(float(v), 1)

    def merge_metrics(keys_scorecol, era_prefix):
        """keys_scorecol: list of ((section,sub), score_label, out_label).
        Merge on Fighter; first key seeds Division/Fights. Blank score cells
        (fighter ranks in one metric's top-30 but not another's) are filled
        from ratings_current.parquet. era_prefix selects the 10yr/5yr column.
        Tracks which (fighter,out_label) cells came from the DB in db_filled."""
        merged = {}
        order = []
        out_labels = []
        for (key, score_label, out_label) in keys_scorecol:
            if out_label not in out_labels:
                out_labels.append(out_label)
            hdr, rows = t[key]
            for r in rows:
                f = r["Fighter"]
                if f not in merged:
                    merged[f] = {
                        "Fighter": f,
                        "Division": r["Division"],
                        "Fights": int(r["Fights"]),
                    }
                    order.append(f)
                merged[f][out_label] = float(r[score_label])
        # fill gaps from DB
        db_filled = set()
        for f in order:
            for out_label in out_labels:
                if merged[f].get(out_label) is None:
                    v = db_score(f, out_label, era_prefix)
                    if v is not None:
                        merged[f][out_label] = v
                        db_filled.add((f, out_label))
        return [merged[f] for f in order], db_filled

    def stack_divisions(section, divisions):
        rows = []
        for div in divisions:
            hdr, drows = t[(section, div)]
            for r in drows:
                rows.append({
                    "Fighter": r["Fighter"],
                    "Division": r["Division"],
                    "Fights": int(r["Fights"]),
                    "WHR Score": float(r["WHR Score"]),
                })
        return rows

    def build_sheet(ws, title, score_cols, data_rows, table_name,
                    divisional=False, db_filled=None):
        """score_cols: list of score-column labels (e.g. ['WHR Score',...]).
        Layout mirrors model: title D2, header D8, data D9+.
        Columns: Fighter, Division, Fights, <scores>, <z-scores>, <ranks>,
        plus Div Rank when divisional."""
        # clear existing content + tables
        for tn in list(ws.tables):
            del ws.tables[tn]
        ws.delete_rows(1, ws.max_row + 1)

        START_COL = 4  # column D
        TITLE_ROW = 2
        HDR_ROW = 8
        DATA_ROW = 9

        z_labels = [f"z {s}" for s in score_cols]
        rank_labels = [f"Rank {s}" for s in score_cols]
        headers = ["Fighter", "Division", "Fights"] + score_cols + z_labels + rank_labels
        if divisional:
            headers += ["Division Rank"]

        # title
        tc = ws.cell(row=TITLE_ROW, column=START_COL, value=title)
        tc.font = Font(size=24)
        tc.alignment = Alignment(horizontal="center")

        # headers
        for j, h in enumerate(headers):
            ws.cell(row=HDR_ROW, column=START_COL + j, value=h)

        n = len(data_rows)
        end_row = DATA_ROW + n - 1

        # column index map (1-based absolute) for formula building
        col_of = {h: START_COL + j for j, h in enumerate(headers)}

        for i, row in enumerate(data_rows):
            r = DATA_ROW + i
            ws.cell(row=r, column=col_of["Fighter"], value=row["Fighter"])
            ws.cell(row=r, column=col_of["Division"], value=row["Division"])
            ws.cell(row=r, column=col_of["Fights"], value=row["Fights"])
            for s in score_cols:
                sc = ws.cell(row=r, column=col_of[s], value=row.get(s))
                # italic flag = value sourced from ratings DB, not markdown top-30
                if db_filled and (row["Fighter"], s) in db_filled:
                    sc.font = Font(italic=True)
                    sc.comment = openpyxl.comments.Comment(
                        "Filled from ratings_current.parquet "
                        "(fighter outside this metric's published top-30)",
                        "build_rankings_sheets",
                    )
            # z-score + rank formulas (structured refs to the table)
            for s in score_cols:
                zc = ws.cell(row=r, column=col_of[f"z {s}"])
                zc.value = (
                    f"=({table_name}[[#This Row],[{s}]]-"
                    f"AVERAGE({table_name}[{s}]))/STDEV({table_name}[{s}])"
                )
                zc.number_format = "0.00"
                rc = ws.cell(row=r, column=col_of[f"Rank {s}"])
                rc.value = (
                    f"=RANK({table_name}[[#This Row],[{s}]],"
                    f"{table_name}[{s}],0)"
                )
            if divisional:
                # rank within division on the (single) score column
                s = score_cols[0]
                dc = ws.cell(row=r, column=col_of["Division Rank"])
                dc.value = (
                    f"=SUMPRODUCT(({table_name}[Division]="
                    f"{table_name}[[#This Row],[Division]])*"
                    f"({table_name}[{s}]>{table_name}[[#This Row],[{s}]]))+1"
                )

        # create the Excel Table
        first = get_column_letter(START_COL)
        last = get_column_letter(START_COL + len(headers) - 1)
        ref = f"{first}{HDR_ROW}:{last}{end_row}"
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tbl)

        # column widths
        widths = {
            "Fighter": 23, "Division": 22, "Fights": 9,
        }
        for s in score_cols:
            widths[s] = 15
        for s in z_labels:
            widths[s] = 11
        for s in rank_labels:
            widths[s] = 14
        if divisional:
            widths["Division Rank"] = 14
        for j, h in enumerate(headers):
            ws.column_dimensions[get_column_letter(START_COL + j)].width = widths[h]

        ws.freeze_panes = ws.cell(row=DATA_ROW, column=START_COL)
        return end_row

    # ---- 1. MEN -WHR 5 yr (3 metrics merged) ---------------------------
    rows, dbf = merge_metrics([
        ((men_overall, "5-Year Peak (WHR Headline)"), "WHR Score", "WHR Score"),
        ((men_overall, "5-Year Peak (WHR + Integrity + Performance)"),
         "WHR I+P Score", "WHR I+P Score"),
        ((men_overall, "5-Year Peak (Method + Integrity + Performance)"),
         "Method I+P Score", "Method I+P Score"),
    ], era_prefix="five_year_peak_headline_mu_")
    build_sheet(
        wb["MEN -WHR 5 yr"], "Men 5-Year Peak Rank",
        ["WHR Score", "WHR I+P Score", "Method I+P Score"],
        rows, "TableMen5yr", db_filled=dbf,
    )

    # ---- 2. MEN -FlyW WHR -> all 8 men divisions stacked ---------------
    men_divs = ["Flyweight", "Bantamweight", "Featherweight", "Lightweight",
                "Welterweight", "Middleweight", "Light Heavyweight", "Heavyweight"]
    rows = stack_divisions(men_div, men_divs)
    ws = wb["MEN -FlyW WHR"]
    end = build_sheet(
        ws, "Men Divisional Rank (WHR 10-Year)",
        ["WHR Score"], rows, "TableMenDiv", divisional=True,
    )

    # ---- 3. WOMEN - Overall -> 10-Year, 3 metrics merged ---------------
    rows, dbf = merge_metrics([
        ((wom_overall, "10-Year Sustained Peak (WHR Headline)"),
         "WHR Score", "WHR Score"),
        ((wom_overall, "10-Year Sustained Peak (WHR + Integrity + Performance)"),
         "WHR I+P Score", "WHR I+P Score"),
        ((wom_overall, "10-Year Sustained Peak (Method + Integrity + Performance)"),
         "Method I+P Score", "Method I+P Score"),
    ], era_prefix="sustained_peak_headline_mu_")
    build_sheet(
        wb["WOMEN - Overall Rankings - 5-Ye"], "Women 10-Year Sustained Peak Rank",
        ["WHR Score", "WHR I+P Score", "Method I+P Score"],
        rows, "TableWomen10yr", db_filled=dbf,
    )
    # rename that sheet to reflect 10-year content
    wb["WOMEN - Overall Rankings - 5-Ye"].title = "WOMEN -Overall 10 yr"

    # ---- 4. NEW: WOMEN 5-Year overall (3 metrics merged) ---------------
    rows, dbf = merge_metrics([
        ((wom_overall, "5-Year Peak (WHR Headline)"), "WHR Score", "WHR Score"),
        ((wom_overall, "5-Year Peak (WHR + Integrity + Performance)"),
         "WHR I+P Score", "WHR I+P Score"),
        ((wom_overall, "5-Year Peak (Method + Integrity + Performance)"),
         "Method I+P Score", "Method I+P Score"),
    ], era_prefix="five_year_peak_headline_mu_")
    ws_new = wb.create_sheet("WOMEN -Overall 5 yr")
    build_sheet(
        ws_new, "Women 5-Year Peak Rank",
        ["WHR Score", "WHR I+P Score", "Method I+P Score"],
        rows, "TableWomen5yr", db_filled=dbf,
    )

    # ---- 5. WOMEN - Divisional -> 3 women divisions stacked ------------
    wom_divs = ["Women's Strawweight", "Women's Flyweight", "Women's Bantamweight"]
    rows = stack_divisions(wom_div, wom_divs)
    build_sheet(
        wb["WOMEN - Divisional Rankings (WH"], "Women Divisional Rank (WHR 10-Year)",
        ["WHR Score"], rows, "TableWomenDiv", divisional=True,
    )
    wb["WOMEN - Divisional Rankings (WH"].title = "WOMEN -Divisional WHR"

    wb.save(XLSX)
    print("Saved. Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
